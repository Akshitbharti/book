import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

# ReportLab for PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import io

# ── Base Book Data ─────────────────────────────────────────────────────────────
BASE_BOOKS = [
    (1,"101","Sarangi (Hindi)",83.00),(2,"102","Mridang (English)",99.00),
    (3,"103","Bhoti Bhasha",60.00),(4,"104","Joyful Mathematics",104.00),
    (5,"201","Sarangi (Hindi)",94.00),(6,"202","Mridang-II (English)",81.00),
    (7,"203","Bhoti Bhasha",60.00),(8,"204","Joyful Mathematics",94.00),
    (9,"301","Veena (Hindi)",103.00),(10,"302","Santoor (English)",104.00),
    (11,"303","Maths Mela (Mathematics)",153.00),(12,"305","Bhoti Bhasha",60.00),
    (13,"306","Our Wondrous World (EVS)",124.00),(14,"401","Veena-II",115.00),
    (15,"402","Santoor-II (English)",98.00),(16,"403","Maths Mela-II (Mathematics)",165.00),
    (17,"405","Bhoti Bhasha",60.00),(18,"406","Our Wondrous World-II (EVS)",126.00),
    (19,"501","Veena-III",109.00),(20,"502","Santoor-III (English)",100.00),
    (21,"503","Maths Mela (Mathematics)",149.00),(22,"507","Our Wondrous World-III (EVS)",139.00),
    (23,"508","Bhoti Bhasha",60.00),(24,"510","Moral Education",16.00),
    (25,"511","History of the Indian Freedom Struggle",16.00),(26,"601","Malhar",90.00),
    (27,"603","Poorvi",87.00),(28,"610","Moral Education",22.00),
    (29,"611","History of the Indian Freedom Struggle",39.00),(30,"612","Deepakam",95.00),
    (31,"613","Ganita Prakash (EM)",197.00),(32,"614","Curiosity",186.00),
    (33,"615","Exploring Society India and Beyond",165.00),(34,"616","Home Science",30.00),
    (35,"623","Hindi Vyakaran und Rachna",90.00),(36,"624","English Grammar",60.00),
    (37,"625","Punjabi Pustak",56.00),(38,"626","Folk Culture of Himachal and Yog",53.00),
    (39,"627","Khayal",83.00),(40,"629","Kriti",134.00),
    (41,"630","Sangeet Pragya (Part-I)",48.00),(42,"701","Vasant-II (Hindi)",70.00),
    (43,"702","Bal Mahabharat-Katha (SR) Hindi",59.00),(44,"703","Honeycomb (English) Core",78.00),
    (45,"704","An Alien Hand (SR) English",39.00),(46,"705","Ganit",189.00),
    (47,"706","Vigyan",148.00),(48,"707","Itihas (Hamare Atit-II) (S.Sc)",98.00),
    (49,"708","Bhugol (Hamara Paryavaran) SSC",52.00),(50,"709","Samajik Aur Rajnitik Jeevan-II (S.Sc)",76.00),
    (51,"710","Ruchira-II (SKT)",59.00),(52,"711","Naitik Shiksha Bhag-VII",24.00),
    (53,"712","Swatantrata Sangram Ka Itihas",32.00),(54,"713","Mathematics",184.00),
    (55,"714","Science",149.00),(56,"715","History (Our Pasts-II) S.Sc",112.00),
    (57,"716","Geography (Our Environment) S.Sc",62.00),(58,"717","Social and Political Life-II, S.Sc",92.00),
    (59,"718","Punjabi",47.00),(60,"719","Himachal Ki Lok Sanskriti Aur Yog",47.00),
    (61,"720","Jaan Pehchan (Urdu)",40.00),(62,"721","Pariyojna Pustika",35.00),
    (63,"801","Vasant Bhag-III",70.00),(64,"802","Bharat Ki Khoj",74.00),
    (65,"803","Honey Dew",65.00),(66,"804","It so Happened (SR)",51.00),
    (67,"805","Ganit",168.00),(68,"806","Vigyan",152.00),
    (69,"807","Hamare Atit-III Bhag-I",89.00),(70,"808","Sansadhan Avam Vikas",52.00),
    (71,"809","Samajik Avam Rajnitik Jeevan-III",104.00),(72,"810","Ruchira Bhag-III",74.00),
    (73,"813","Naitik Shiksha",20.00),(74,"814","Swatantrata Sangram Ka Itihas",39.00),
    (75,"815","Mathematics (Eng. Medium)",164.00),(76,"816","Science",160.00),
    (77,"817","Our Pasts-III (New Edition)",98.00),(78,"818","Social and Political Life-III",109.00),
    (79,"819","Resources and Development",55.00),(80,"820","Punjabi",71.00),
    (81,"821","Hamare Atit-III Bhag-II",27.00),(82,"822","Our Pasts-III (Bhag-II)",26.00),
    (83,"823","Himachal Ki Lok Sanskriti Aur Yog",55.00),(84,"824","Jaan Pahchan (Urdu)",33.00),
    (85,"825","Pariyojna Pustika",33.00),(86,"901","Kshitij (Hindi)",65.00),
    (87,"902","Kritika (SR) Hindi",33.00),(88,"903","Beehive (English)",72.00),
    (89,"904","Moments (SR) English",40.00),(90,"905","Ganit",170.00),
    (91,"906","Vigyan",152.00),(92,"907","Bharat Aur Samkalin Vishwa-I S.Sc",115.00),
    (93,"908","Samkalin Bharat-I (Bhugol) S.Sc",50.00),(94,"909","Loktantrik Rajniti (PSc) S.Sc",80.00),
    (95,"910","Nawa-e-Urdu",94.00),(96,"911","Naitik Shiksha Bhag-I",36.00),
    (97,"912","Swatantrata Sangram Ka Itihas",45.00),(98,"913","Shemushi Prathma Bhag",65.00),
    (99,"914","Punjabi Book-I",93.00),(100,"915","Arthshashtra",45.00),
    (101,"916","Mathematics (Eng. Medium)",149.00),(102,"917","Science (Eng. Medium)",137.00),
    (103,"918","Aapda Prabandhan (S.Sc)",61.00),(104,"922","Kala Sanklan",118.00),
    (105,"925","Science (Practical Book)",103.00),(106,"926","Grih Vigyan",106.00),
    (107,"929","Vanijya Pranali Avam Vahi Khata",23.00),(108,"930","Hamari Arthvyavastha ka Parichaya",32.00),
    (109,"932","Sanskrit Vyakaran Kaumudi",52.00),(110,"933","English Grammar",102.00),
    (111,"934","Shiksharthi Vyakaran Aur Vyavahari Hindi",78.00),(112,"951","Disaster Management (SSc)",60.00),
    (113,"952","Contemporary India-I SSc.",57.00),(114,"953","Economics",56.00),
    (115,"1135","Statistics for Economics (English Medium)",89.00),
    (116,"1136","Indian Economic Development (English Medium)",132.00),
    (117,"1137","Practical work in geography Part-I",106.00),
    (118,"1201","Antra-2 (Hindi)",86.00),(119,"1202","Antral-2 Suppl. (Hindi)",32.00),
    (120,"1203","Flamingo-English Core",63.00),(121,"1204","Vistas, Suppl. English Core",45.00),
    (122,"1205","Shaswati-II (Sanskrit)",81.00),(123,"1206","Mathematics Part-I",136.00),
    (124,"1207","Mathematics Part-II",143.00),(125,"1213","Bhartiya Itihas ke kuch Vishay Bhag-I",80.00),
    (126,"1214","Bhartiya Itihas Ke Kuch Vishay Bhag-II",102.00),
    (127,"1215","Bhartiya Itihas Ke Kuch Vishay Bhag-III",115.00),
    (128,"1218","Samkalin Vishwa Rajniti (PSc)",101.00),(129,"1219","Swatantar Bharat Mein Rajniti (PSc)",135.00),
    (130,"1220","Biology",184.00),(131,"1221","Physics-I",170.00),
    (132,"1222","Physics-II",140.00),(133,"1223","Chemistry-I",170.00),
    (134,"1224","Chemistry-II",128.00),(135,"1225","Biology Practical Book",155.00),
    (136,"1226","Physics Practical Book",195.00),(137,"1227","Chemistry Practical Book",180.00),
    (138,"1228","Computer Science 12th",210.00),(139,"1229","Computer Science 12th Practical book",180.00),
    (140,"1230","Sharirik Shiksha (Hindi Medium)",85.00),(141,"1231","Physical Education (English Medium)",110.00),
    (142,"1232","Gulistan-e-Adab",97.00),(143,"1233","Accountancy-I (EM)",109.00),
    (144,"1234","Lekhashastra Sajhedari Khate Bhag-I (HM)",115.00),
    (145,"1235","Accountancy-II Company Accounts (EM)",155.00),
    (146,"1236","Lekhashastra-Vittiya Lekhaunkan Bhag-II (HM)",135.00),
    (147,"1237","Business Studies-II (English Medium)",61.00),(148,"1238","Vyavsay Adhyayan-II (Hindi Medium)",46.00),
    (149,"1239","Introductory Microeconomics (English Medium)",73.00),
    (150,"1240","Vyashti Arthshastra ek Parichay (Hindi Medium)",79.00),
    (151,"1241","Introductory Macroeconomics (English Medium)",82.00),
    (152,"1242","Samashti Arthshastra ek Parichay (Hindi Medium)",87.00),
    (153,"1243","Practical work in geography Part-II",56.00),
]

# ── Excel Generator ────────────────────────────────────────────────────────────
def generate_excel_bytes(summary_rows, grand_total):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Book Order"

    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Book Order Summary"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", start_color="0A3D62")
    c.alignment = center
    ws.row_dimensions[1].height = 30

    headers    = ["Sr. No.", "Book Code", "Book Name", "Unit Price (Rs)", "Quantity", "Subtotal (Rs)"]
    col_widths = [8, 12, 45, 18, 12, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", start_color="1565C0")
        c.alignment = center
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    alt_fill = PatternFill("solid", start_color="E3F2FD")
    for i, row_data in enumerate(summary_rows, 1):
        row  = i + 2
        fill = alt_fill if i % 2 == 0 else None
        price_val    = float(str(row_data["Unit Price (Rs)"]).replace("Rs", "").replace(",", "").strip())
        subtotal_val = float(str(row_data["Subtotal (Rs)"]).replace("Rs", "").replace(",", "").strip())
        vals = [i, row_data["Book Code"], row_data["Book Name"], price_val, row_data["Quantity"], subtotal_val]
        alns = [center, center, left, center, center, center]
        for col, (val, aln) in enumerate(zip(vals, alns), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.alignment = aln
            c.border    = border
            if fill:
                c.fill = fill
        ws.row_dimensions[row].height = 18

    tr = len(summary_rows) + 3
    ws.merge_cells(f"A{tr}:E{tr}")
    lbl           = ws.cell(row=tr, column=1, value="GRAND TOTAL")
    lbl.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    lbl.fill      = PatternFill("solid", start_color="0D47A1")
    lbl.alignment = right
    lbl.border    = border

    tc            = ws.cell(row=tr, column=6, value=grand_total)
    tc.font       = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.fill       = PatternFill("solid", start_color="0D47A1")
    tc.alignment  = center
    tc.border     = border
    ws.row_dimensions[tr].height = 22
    ws.freeze_panes = "A3"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    with open(tmp.name, "rb") as f:
        data = f.read()
    os.unlink(tmp.name)
    return data


# ── PDF Generator ──────────────────────────────────────────────────────────────
def generate_pdf_bytes(summary_rows, grand_total):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    # Colours
    dark_blue  = colors.HexColor("#0A3D62")
    mid_blue   = colors.HexColor("#1565C0")
    navy       = colors.HexColor("#0D47A1")
    light_blue = colors.HexColor("#E3F2FD")
    white      = colors.white
    black      = colors.black

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "title", parent=styles["Normal"],
        fontSize=16, fontName="Helvetica-Bold",
        textColor=white, alignment=TA_CENTER,
        spaceAfter=0,
    )
    normal_center = ParagraphStyle(
        "nc", parent=styles["Normal"],
        fontSize=8, alignment=TA_CENTER,
    )
    normal_left = ParagraphStyle(
        "nl", parent=styles["Normal"],
        fontSize=8, alignment=TA_LEFT,
    )

    story = []

    # ── Title block ──
    title_data = [[Paragraph("Book Order Summary", title_style)]]
    title_tbl  = Table(title_data, colWidths=[17.7*cm])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), dark_blue),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ("RIGHTPADDING",  (0,0), (-1,-1), 8),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 0.3*cm))

    # ── Table header ──
    headers = ["Sr.\nNo.", "Book\nCode", "Book Name", "Unit Price\n(Rs)", "Qty", "Subtotal\n(Rs)"]
    col_w   = [1.2*cm, 2*cm, 8.5*cm, 2.5*cm, 1.5*cm, 2*cm]

    header_row = [Paragraph(h, ParagraphStyle(
        "hdr", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica-Bold",
        textColor=white, alignment=TA_CENTER,
    )) for h in headers]

    # ── Data rows ──
    data = [header_row]
    for i, row_data in enumerate(summary_rows, 1):
        price_val    = float(str(row_data["Unit Price (Rs)"]).replace("Rs", "").replace(",", "").strip())
        subtotal_val = float(str(row_data["Subtotal (Rs)"]).replace("Rs", "").replace(",", "").strip())
        row = [
            Paragraph(str(i),                         normal_center),
            Paragraph(str(row_data["Book Code"]),     normal_center),
            Paragraph(str(row_data["Book Name"]),     normal_left),
            Paragraph(f"Rs {price_val:.2f}",          normal_center),
            Paragraph(str(row_data["Quantity"]),      normal_center),
            Paragraph(f"Rs {subtotal_val:.2f}",       normal_center),
        ]
        data.append(row)

    # Grand total row
    grand_style = ParagraphStyle(
        "gt", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=white, alignment=TA_RIGHT,
    )
    grand_val_style = ParagraphStyle(
        "gtv", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=white, alignment=TA_CENTER,
    )
    data.append([
        Paragraph("GRAND TOTAL", grand_style),
        "", "", "", "",
        Paragraph(f"Rs {grand_total:.2f}", grand_val_style),
    ])

    tbl = Table(data, colWidths=col_w, repeatRows=1)

    # Build alternating row styles
    style_cmds = [
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  mid_blue),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  white),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",        (0, 0), (-1,-1),  "MIDDLE"),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        ("TOPPADDING",    (0, 0), (-1, 0),  6),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  6),
        # Grid
        ("GRID",          (0, 0), (-1,-2),  0.4, colors.HexColor("#BBBBBB")),
        ("TOPPADDING",    (0, 1), (-1,-2),  4),
        ("BOTTOMPADDING", (0, 1), (-1,-2),  4),
        ("LEFTPADDING",   (0, 0), (-1,-1),  4),
        ("RIGHTPADDING",  (0, 0), (-1,-1),  4),
        # Grand total row
        ("BACKGROUND",   (0,-1), (-1,-1), navy),
        ("SPAN",         (0,-1), (4,-1)),
        ("TOPPADDING",   (0,-1), (-1,-1), 6),
        ("BOTTOMPADDING",(0,-1), (-1,-1), 6),
        ("LINEABOVE",    (0,-1), (-1,-1), 1, navy),
    ]

    # Alternating row colours for data rows
    for i in range(1, len(summary_rows) + 1):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), light_blue))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Book Order System", layout="wide")
st.title("📚 Book Order System")
st.markdown("Select books and set quantity — total price updates automatically.")

# ── Session State Init ─────────────────────────────────────────────────────────
if "quantities"   not in st.session_state: st.session_state.quantities   = {}
if "custom_books" not in st.session_state: st.session_state.custom_books = []

# ── Build full dataframe (base + custom) ───────────────────────────────────────
all_books = list(BASE_BOOKS)
start_sr  = len(BASE_BOOKS) + 1
for i, cb in enumerate(st.session_state.custom_books):
    all_books.append((start_sr + i, cb["code"], cb["name"], cb["price"]))

df = pd.DataFrame(all_books, columns=["Sr. No.", "Book Code", "Book Name", "Price (Rs)"])

# ── Add Custom Book Section ────────────────────────────────────────────────────
with st.expander("➕ Add a New Book to the List", expanded=False):
    st.markdown("Fill in the details below and click **Add Book**.")
    col1, col2, col3, col4 = st.columns([1.5, 3, 1.5, 1])
    new_code  = col1.text_input("Book Code",  placeholder="e.g. 9999",       key="new_code")
    new_name  = col2.text_input("Book Name",  placeholder="e.g. My New Book", key="new_name")
    new_price = col3.number_input("Price (Rs)", min_value=0.0, step=0.5, format="%.2f", key="new_price")
    add_clicked = col4.button("✅ Add Book", use_container_width=True)

    if add_clicked:
        if not new_code.strip():
            st.error("❌ Book Code cannot be empty.")
        elif not new_name.strip():
            st.error("❌ Book Name cannot be empty.")
        elif new_price <= 0:
            st.error("❌ Price must be greater than 0.")
        elif new_code.strip() in df["Book Code"].values:
            st.error(f"❌ Book Code **{new_code.strip()}** already exists.")
        else:
            st.session_state.custom_books.append({
                "code":  new_code.strip(),
                "name":  new_name.strip(),
                "price": float(new_price),
            })
            st.success(f"✅ Book **{new_name.strip()}** added successfully!")
            st.rerun()

    if st.session_state.custom_books:
        st.markdown("**Custom books added this session:**")
        for idx, cb in enumerate(st.session_state.custom_books):
            cc1, cc2, cc3, cc4 = st.columns([1.5, 3, 1.5, 1])
            cc1.write(cb["code"])
            cc2.write(cb["name"])
            cc3.write(f"Rs {cb['price']:.2f}")
            if cc4.button("🗑️", key=f"del_custom_{idx}", help="Remove this book"):
                st.session_state.quantities.pop(cb["code"], None)
                st.session_state.custom_books.pop(idx)
                st.rerun()

st.divider()

# Rebuild df after custom book changes
all_books = list(BASE_BOOKS)
for i, cb in enumerate(st.session_state.custom_books):
    all_books.append((start_sr + i, cb["code"], cb["name"], cb["price"]))
df = pd.DataFrame(all_books, columns=["Sr. No.", "Book Code", "Book Name", "Price (Rs)"])

# ── Search / Filter ────────────────────────────────────────────────────────────
search = st.text_input("🔍 Search books by name or code", placeholder="e.g. Mathematics, 1228, Science…")

filtered_df = df.copy()
if search.strip():
    q = search.strip().lower()
    filtered_df = df[
        df["Book Name"].str.lower().str.contains(q) |
        df["Book Code"].str.lower().str.contains(q)
    ]

st.markdown(f"**{len(filtered_df)} book(s) shown**")

# ── Book Selection Table ───────────────────────────────────────────────────────
st.subheader("📋 Select Books")

cols = st.columns([0.5, 1.2, 4, 1.5, 1.5, 1.5])
for h, c in zip(["Select", "Code", "Book Name", "Price (Rs)", "Qty", "Subtotal (Rs)"], cols):
    c.markdown(f"**{h}**")
st.divider()

for _, row in filtered_df.iterrows():
    code  = row["Book Code"]
    name  = row["Book Name"]
    price = row["Price (Rs)"]
    qty   = st.session_state.quantities.get(code, 0)

    c1, c2, c3, c4, c5, c6 = st.columns([0.5, 1.2, 4, 1.5, 1.5, 1.5])

    selected = c1.checkbox("Select", value=(qty > 0), key=f"chk_{code}", label_visibility="collapsed")
    c2.write(code)
    c3.write(name)
    c4.write(f"Rs {price:.2f}")

    if selected:
        new_qty = c5.number_input(
            "Qty", min_value=1, max_value=999,
            value=max(qty, 1),
            key=f"qty_{code}", label_visibility="collapsed"
        )
        st.session_state.quantities[code] = new_qty
        c6.write(f"Rs {price * new_qty:.2f}")
    else:
        st.session_state.quantities[code] = 0
        c5.write("—")
        c6.write("—")

# ── Order Summary ──────────────────────────────────────────────────────────────
ordered = {code: qty for code, qty in st.session_state.quantities.items() if qty > 0}

st.divider()
st.subheader("🛒 Order Summary")

if not ordered:
    st.info("No books selected yet. Check the boxes above to add books to your order.")
else:
    summary_rows = []
    grand_total  = 0.0

    for code, qty in ordered.items():
        match = df[df["Book Code"] == code]
        if match.empty:
            continue
        book_row = match.iloc[0]
        price    = book_row["Price (Rs)"]
        subtotal = price * qty
        grand_total += subtotal
        summary_rows.append({
            "Book Code":       code,
            "Book Name":       book_row["Book Name"],
            "Unit Price (Rs)": f"Rs {price:.2f}",
            "Quantity":        qty,
            "Subtotal (Rs)":   f"Rs {subtotal:.2f}",
        })

    summary_df = pd.DataFrame(summary_rows)
    st.dataframe(summary_df, width="stretch", hide_index=True)

    st.markdown(
        f"""
        <div style="text-align:right; font-size:1.4rem; font-weight:700;
                    background:#f0f7ff; padding:12px 20px; border-radius:8px;
                    border-left:5px solid #1976d2; margin-top:8px;">
            💰 Grand Total: &nbsp; Rs {grand_total:.2f}
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("&nbsp;")

    # ── Cache Excel & PDF bytes ────────────────────────────────────────────────
    order_key = str(sorted(ordered.items()))
    if st.session_state.get("excel_key") != order_key:
        st.session_state.excel_bytes = generate_excel_bytes(summary_rows, grand_total)
        st.session_state.pdf_bytes   = generate_pdf_bytes(summary_rows, grand_total)
        st.session_state.excel_key   = order_key

    # ── Download Buttons side by side ─────────────────────────────────────────
    dl1, dl2 = st.columns(2)

    with dl1:
        st.download_button(
            label="📥 Download Order as Excel",
            data=st.session_state.excel_bytes,
            file_name="Book_Order_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

    with dl2:
        st.download_button(
            label="📄 Download Order as PDF",
            data=st.session_state.pdf_bytes,
            file_name="Book_Order_Summary.pdf",
            mime="application/pdf",
            type="primary",
            use_container_width=True,
        )

    st.markdown("&nbsp;")

    if st.button("🗑️ Clear Order"):
        st.session_state.quantities = {}
        st.rerun()